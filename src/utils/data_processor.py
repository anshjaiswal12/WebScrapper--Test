"""
Data processing module for cleaning, normalizing, and exporting scraped data.

This module provides comprehensive data manipulation functions including:
- Data combination and merging
- Cleaning and normalization
- Price comparison calculations
- Export to CSV and Excel with formatting
- Summary report generation
"""
import pandas as pd
import numpy as np
from typing import List, Dict, Optional, Tuple, Any
from datetime import datetime
import os
import re

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import CellIsRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel formatting will be limited")

from src.utils.logger import setup_logger
from config.settings import CSV_OUTPUT_DIR, EXCEL_OUTPUT_DIR

logger = setup_logger()


class DataProcessor:
    """
    Process and organize scraped product data.
    
    Provides methods for combining, cleaning, analyzing, and exporting
    product data from multiple sources.
    """
    
    def __init__(self):
        """Initialize data processor."""
        self.df: Optional[pd.DataFrame] = None
        logger.info("DataProcessor initialized")
    
    def combine_scraped_data(self, amazon_data: List[Dict], flipkart_data: List[Dict]) -> pd.DataFrame:
        """
        Merge data from all sources into single DataFrame.
        
        Combines Amazon and Flipkart product lists, handles duplicates,
        adds timestamp, and sorts by price (lowest first).
        
        Args:
            amazon_data: List of product dictionaries from Amazon
            flipkart_data: List of product dictionaries from Flipkart
            
        Returns:
            Combined and sorted DataFrame
            
        Raises:
            ValueError: If both data lists are empty
        """
        try:
            # Combine all data
            all_products = []
            
            if amazon_data:
                all_products.extend(amazon_data)
                logger.info(f"Added {len(amazon_data)} products from Amazon")
            
            if flipkart_data:
                all_products.extend(flipkart_data)
                logger.info(f"Added {len(flipkart_data)} products from Flipkart")
            
            if not all_products:
                logger.warning("No products to combine - both data sources are empty")
                return pd.DataFrame()
            
            # Convert to DataFrame
            df = pd.DataFrame(all_products)
            
            # Normalize column names (handle both 'title' and 'product_name')
            if 'product_name' in df.columns and 'title' not in df.columns:
                df['title'] = df['product_name']
            elif 'title' in df.columns and 'product_name' not in df.columns:
                df['product_name'] = df['title']
            
            # Ensure required columns exist
            required_columns = {
                'title': None,
                'product_name': None,
                'price': None,
                'source': None
            }
            
            for col, default in required_columns.items():
                if col not in df.columns:
                    df[col] = default
                    logger.debug(f"Added missing column: {col}")
            
            # Fill missing optional columns
            optional_columns = ['rating', 'url', 'sponsored', 'reviews_count', 'image_url']
            for col in optional_columns:
                if col not in df.columns:
                    df[col] = None
            
            # Add timestamp
            df['scraped_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Handle duplicate products
            # Identify duplicates based on product name similarity and price proximity
            df = self._handle_duplicates(df)
            
            # Sort by price (lowest first)
            if 'price' in df.columns and df['price'].notna().any():
                df = df.sort_values(by='price', ascending=True, na_position='last')
                logger.info("Sorted data by price (lowest first)")
            
            self.df = df
            logger.info(f"✓ Successfully combined {len(df)} products from {len(amazon_data)} Amazon + {len(flipkart_data)} Flipkart sources")
            
            return df
            
        except Exception as e:
            logger.error(f"Error combining scraped data: {e}", exc_info=True)
            return pd.DataFrame()
    
    def _handle_duplicates(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Handle duplicate products across sources.
        
        Args:
            df: DataFrame with potential duplicates
            
        Returns:
            DataFrame with duplicates handled
        """
        if df.empty:
            return df
        
        # Simple duplicate detection based on exact name and similar price
        # Keep the one with lower price or better rating
        initial_count = len(df)
        
        # Remove exact duplicates (same name, source, and price)
        df = df.drop_duplicates(subset=['product_name', 'source', 'price'], keep='first')
        
        # For similar products (same name, different source), keep the one with lower price
        if 'product_name' in df.columns and 'price' in df.columns:
            df = df.sort_values(by=['product_name', 'price'], ascending=[True, True])
            df = df.drop_duplicates(subset=['product_name'], keep='first')
        
        removed = initial_count - len(df)
        if removed > 0:
            logger.info(f"Removed {removed} duplicate products")
        
        return df
    
    def calculate_price_statistics(self, df: Optional[pd.DataFrame] = None) -> Dict[str, Any]:
        """
        Calculate comprehensive price comparison metrics.
        
        Calculates:
        - Lowest price and source
        - Highest price and source
        - Average price
        - Price difference (max - min)
        - Percentage savings
        - Best rated product
        
        Args:
            df: DataFrame to analyze (uses self.df if None)
            
        Returns:
            Dictionary with statistics
        """
        df_to_analyze = df if df is not None else self.df
        
        if df_to_analyze is None or df_to_analyze.empty:
            logger.warning("No data available for statistics calculation")
            return {}
        
        if 'price' not in df_to_analyze.columns:
            logger.warning("Price column not found in DataFrame")
            return {}
        
        stats = {}
        
        try:
            # Filter out invalid prices
            valid_prices = df_to_analyze[df_to_analyze['price'].notna() & (df_to_analyze['price'] > 0)]
            
            if valid_prices.empty:
                logger.warning("No valid prices found for statistics")
                return {}
            
            # Lowest price and source
            lowest_idx = valid_prices['price'].idxmin()
            lowest_product = valid_prices.loc[lowest_idx]
            stats['lowest_price'] = {
                'price': float(lowest_product['price']),
                'product_name': str(lowest_product.get('product_name', lowest_product.get('title', 'Unknown'))),
                'source': str(lowest_product.get('source', 'Unknown')),
                'url': str(lowest_product.get('url', ''))
            }
            
            # Highest price and source
            highest_idx = valid_prices['price'].idxmax()
            highest_product = valid_prices.loc[highest_idx]
            stats['highest_price'] = {
                'price': float(highest_product['price']),
                'product_name': str(highest_product.get('product_name', highest_product.get('title', 'Unknown'))),
                'source': str(highest_product.get('source', 'Unknown')),
                'url': str(highest_product.get('url', ''))
            }
            
            # Average price
            stats['average_price'] = float(valid_prices['price'].mean())
            stats['median_price'] = float(valid_prices['price'].median())
            
            # Price difference
            price_range = stats['highest_price']['price'] - stats['lowest_price']['price']
            stats['price_difference'] = price_range
            
            # Percentage savings (compared to highest)
            if stats['highest_price']['price'] > 0:
                savings_percent = (price_range / stats['highest_price']['price']) * 100
                stats['percentage_savings'] = round(savings_percent, 2)
            else:
                stats['percentage_savings'] = 0.0
            
            # Best rated product
            if 'rating' in df_to_analyze.columns:
                valid_ratings = df_to_analyze[
                    df_to_analyze['rating'].notna() & (df_to_analyze['rating'] > 0)
                ]
                if not valid_ratings.empty:
                    best_rated_idx = valid_ratings['rating'].idxmax()
                    best_rated = valid_ratings.loc[best_rated_idx]
                    stats['best_rated'] = {
                        'rating': float(best_rated['rating']),
                        'product_name': str(best_rated.get('product_name', best_rated.get('title', 'Unknown'))),
                        'source': str(best_rated.get('source', 'Unknown')),
                        'price': float(best_rated.get('price', 0)),
                        'url': str(best_rated.get('url', ''))
                    }
            
            # Source-wise statistics
            if 'source' in df_to_analyze.columns:
                source_stats = {}
                for source in df_to_analyze['source'].unique():
                    source_df = df_to_analyze[df_to_analyze['source'] == source]
                    source_prices = source_df[source_df['price'].notna() & (source_df['price'] > 0)]['price']
                    if not source_prices.empty:
                        source_stats[source] = {
                            'count': len(source_df),
                            'min_price': float(source_prices.min()),
                            'max_price': float(source_prices.max()),
                            'avg_price': float(source_prices.mean())
                        }
                stats['source_statistics'] = source_stats
            
            logger.info("✓ Price statistics calculated successfully")
            return stats
            
        except Exception as e:
            logger.error(f"Error calculating price statistics: {e}", exc_info=True)
            return {}
    
    def clean_product_names(self, df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
        """
        Standardize product names for comparison.
        
        Performs:
        - Remove extra whitespace
        - Standardize brand names (HP vs hp)
        - Extract model numbers
        - Group similar products
        
        Args:
            df: DataFrame to clean (uses self.df if None)
            
        Returns:
            Cleaned DataFrame
        """
        df_to_clean = df.copy() if df is not None else (self.df.copy() if self.df is not None else pd.DataFrame())
        
        if df_to_clean.empty:
            logger.warning("No data to clean")
            return df_to_clean
        
        # Determine which name column to use
        name_col = None
        for col in ['product_name', 'title']:
            if col in df_to_clean.columns:
                name_col = col
                break
        
        if name_col is None:
            logger.warning("No product name column found")
            return df_to_clean
        
        try:
            # Remove extra whitespace
            df_to_clean[name_col] = df_to_clean[name_col].astype(str).str.strip()
            df_to_clean[name_col] = df_to_clean[name_col].str.replace(r'\s+', ' ', regex=True)
            
            # Standardize brand names (case-insensitive)
            brand_replacements = {
                r'\bhp\b': 'HP',
                r'\bhp\s+': 'HP ',
                r'\bpavilion\b': 'Pavilion',
                r'\bPavilion\b': 'Pavilion'
            }
            
            for pattern, replacement in brand_replacements.items():
                df_to_clean[name_col] = df_to_clean[name_col].str.replace(
                    pattern, replacement, regex=True, case=False
                )
            
            # Extract model numbers (common patterns: numbers, letters, dashes)
            if 'model_number' not in df_to_clean.columns:
                df_to_clean['model_number'] = df_to_clean[name_col].str.extract(
                    r'([A-Z]{1,3}[-]?\d{3,5}[A-Z]?)', expand=False
                )
            
            # Group similar products (normalize for comparison)
            if 'normalized_name' not in df_to_clean.columns:
                df_to_clean['normalized_name'] = df_to_clean[name_col].str.lower()
                df_to_clean['normalized_name'] = df_to_clean['normalized_name'].str.replace(
                    r'[^\w\s]', '', regex=True
                )
                df_to_clean['normalized_name'] = df_to_clean['normalized_name'].str.replace(
                    r'\s+', ' ', regex=True
                )
            
            logger.info(f"✓ Cleaned {len(df_to_clean)} product names")
            
            if df is None:
                self.df = df_to_clean
            
            return df_to_clean
            
        except Exception as e:
            logger.error(f"Error cleaning product names: {e}", exc_info=True)
            return df_to_clean
    
    def export_to_csv(self, filename: str, df: Optional[pd.DataFrame] = None) -> Tuple[bool, str]:
        """
        Export to CSV with proper formatting.
        
        Saves to data/output/ directory with timestamp in filename.
        Uses UTF-8 encoding for proper character support.
        
        Args:
            filename: Output filename (timestamp will be added if not present)
            df: DataFrame to export (uses self.df if None)
            
        Returns:
            Tuple of (success: bool, message: str)
        """
        df_to_export = df if df is not None else self.df
        
        if df_to_export is None or df_to_export.empty:
            msg = "✗ Export failed: No data to export"
            logger.warning(msg)
            return False, msg
        
        try:
            # Ensure filename has timestamp if not present
            if 'price_comparison_' not in filename or '.csv' not in filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                base_name = filename.replace('.csv', '') if filename.endswith('.csv') else filename
                filename = f"{base_name}_{timestamp}.csv"
            
            # Ensure filename ends with .csv
            if not filename.endswith('.csv'):
                filename += '.csv'
            
            # Build full path
            if not os.path.isabs(filename):
                filepath = os.path.join(CSV_OUTPUT_DIR, filename)
            else:
                filepath = filename
            
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            # Export to CSV with UTF-8 encoding
            df_to_export.to_csv(filepath, index=False, encoding='utf-8')
            
            msg = f"✓ Successfully exported {len(df_to_export)} products to {filepath}"
            logger.info(msg)
            return True, msg
            
        except PermissionError:
            msg = f"✗ Export failed: Permission denied for {filename}"
            logger.error(msg)
            return False, msg
        except Exception as e:
            msg = f"✗ Export failed: {str(e)}"
            logger.error(msg, exc_info=True)
            return False, msg
    
    def export_to_excel(self, filename: str, df: Optional[pd.DataFrame] = None) -> Tuple[bool, str]:
        """
        Export to Excel with comprehensive formatting.
        
        Creates multiple sheets:
        - Raw Data: All product information
        - Summary: Key statistics
        - Statistics: Detailed price comparison metrics
        
        Formatting includes:
        - Auto-fit column widths
        - Bold headers
        - Conditional formatting (highlight lowest price in green)
        - Filters on headers
        
        Args:
            filename: Output filename (timestamp will be added if not present)
            df: DataFrame to export (uses self.df if None)
            
        Returns:
            Tuple of (success: bool, message: str)
        """
        df_to_export = df if df is not None else self.df
        
        if df_to_export is None or df_to_export.empty:
            msg = "✗ Export failed: No data to export"
            logger.warning(msg)
            return False, msg
        
        if not OPENPYXL_AVAILABLE:
            msg = "✗ Export failed: openpyxl not available for Excel formatting"
            logger.error(msg)
            return False, msg
        
        try:
            # Ensure filename has timestamp if not present
            if 'price_comparison_' not in filename or '.xlsx' not in filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                base_name = filename.replace('.xlsx', '') if filename.endswith('.xlsx') else filename
                filename = f"{base_name}_{timestamp}.xlsx"
            
            # Ensure filename ends with .xlsx
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            # Build full path
            if not os.path.isabs(filename):
                filepath = os.path.join(EXCEL_OUTPUT_DIR, filename)
            else:
                filepath = filename
            
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            # Calculate statistics
            stats = self.calculate_price_statistics(df_to_export)
            
            # Write to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Sheet 1: Raw Data
                df_to_export.to_excel(writer, sheet_name='Raw Data', index=False)
                
                # Sheet 2: Summary
                summary_data = {
                    'Metric': [
                        'Total Products',
                        'Lowest Price',
                        'Highest Price',
                        'Average Price',
                        'Price Difference',
                        'Percentage Savings',
                        'Best Rated Product'
                    ],
                    'Value': [
                        len(df_to_export),
                        f"₹{stats.get('lowest_price', {}).get('price', 0):,.0f}" if stats.get('lowest_price') else 'N/A',
                        f"₹{stats.get('highest_price', {}).get('price', 0):,.0f}" if stats.get('highest_price') else 'N/A',
                        f"₹{stats.get('average_price', 0):,.0f}" if stats.get('average_price') else 'N/A',
                        f"₹{stats.get('price_difference', 0):,.0f}" if stats.get('price_difference') else 'N/A',
                        f"{stats.get('percentage_savings', 0):.2f}%" if stats.get('percentage_savings') else 'N/A',
                        stats.get('best_rated', {}).get('product_name', 'N/A') if stats.get('best_rated') else 'N/A'
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Sheet 3: Statistics
                if stats:
                    stats_list = []
                    for key, value in stats.items():
                        if isinstance(value, dict):
                            for sub_key, sub_value in value.items():
                                stats_list.append({
                                    'Category': key,
                                    'Metric': sub_key,
                                    'Value': str(sub_value)
                                })
                        else:
                            stats_list.append({
                                'Category': 'General',
                                'Metric': key,
                                'Value': str(value)
                            })
                    stats_df = pd.DataFrame(stats_list)
                    stats_df.to_excel(writer, sheet_name='Statistics', index=False)
            
            # Apply formatting
            self._format_excel_file(filepath, df_to_export, stats)
            
            msg = f"✓ Successfully exported {len(df_to_export)} products to {filepath}"
            logger.info(msg)
            return True, msg
            
        except PermissionError:
            msg = f"✗ Export failed: Permission denied for {filename}"
            logger.error(msg)
            return False, msg
        except Exception as e:
            msg = f"✗ Export failed: {str(e)}"
            logger.error(msg, exc_info=True)
            return False, msg
    
    def _format_excel_file(self, filepath: str, df: pd.DataFrame, stats: Dict):
        """
        Apply formatting to Excel file.
        
        Args:
            filepath: Path to Excel file
            df: DataFrame with data
            stats: Statistics dictionary
        """
        if not OPENPYXL_AVAILABLE:
            return
        
        try:
            wb = load_workbook(filepath)
            
            # Format Raw Data sheet
            if 'Raw Data' in wb.sheetnames:
                ws = wb['Raw Data']
                
                # Bold headers
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-fit column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Add filters
                ws.auto_filter.ref = ws.dimensions
                
                # Conditional formatting for lowest price (green)
                if 'price' in df.columns and stats.get('lowest_price'):
                    price_col = None
                    for idx, col in enumerate(df.columns, start=1):
                        if col == 'price':
                            price_col = idx
                            break
                    
                    if price_col:
                        lowest_price = stats['lowest_price']['price']
                        # Apply green fill to cells with lowest price
                        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        for row in range(2, len(df) + 2):
                            cell = ws.cell(row=row, column=price_col)
                            if cell.value == lowest_price:
                                cell.fill = green_fill
            
            # Format Summary sheet
            if 'Summary' in wb.sheetnames:
                ws = wb['Summary']
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                # Auto-fit columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
            logger.debug("Applied Excel formatting")
                
        except Exception as e:
            logger.warning(f"Could not apply Excel formatting: {e}")
    
    def generate_summary_report(self, df: Optional[pd.DataFrame] = None) -> str:
        """
        Create text summary of findings.
        
        Includes:
        - Total products scraped
        - Price range
        - Best deal recommendation
        - Average rating by source
        
        Args:
            df: DataFrame to analyze (uses self.df if None)
            
        Returns:
            Formatted text report
        """
        df_to_analyze = df if df is not None else self.df
        
        if df_to_analyze is None or df_to_analyze.empty:
            return "No data available for summary report."
        
        try:
            stats = self.calculate_price_statistics(df_to_analyze)
            
            report_lines = []
            report_lines.append("=" * 70)
            report_lines.append("PRICE COMPARISON SUMMARY REPORT")
            report_lines.append("=" * 70)
            report_lines.append("")
            
            # Total products
            report_lines.append(f"Total Products Scraped: {len(df_to_analyze)}")
            
            # Source breakdown
            if 'source' in df_to_analyze.columns:
                source_counts = df_to_analyze['source'].value_counts()
                report_lines.append("\nProducts by Source:")
                for source, count in source_counts.items():
                    report_lines.append(f"  - {source}: {count} products")
            
            # Price range
            if stats.get('lowest_price') and stats.get('highest_price'):
                report_lines.append("\nPrice Range:")
                report_lines.append(f"  Lowest:  ₹{stats['lowest_price']['price']:,.0f} ({stats['lowest_price']['source']})")
                report_lines.append(f"  Highest: ₹{stats['highest_price']['price']:,.0f} ({stats['highest_price']['source']})")
                report_lines.append(f"  Average: ₹{stats.get('average_price', 0):,.0f}")
                report_lines.append(f"  Difference: ₹{stats.get('price_difference', 0):,.0f} ({stats.get('percentage_savings', 0):.2f}% savings)")
            
            # Best deal recommendation
            if stats.get('lowest_price'):
                lowest = stats['lowest_price']
                report_lines.append("\n🏆 Best Deal Recommendation:")
                report_lines.append(f"  Product: {lowest['product_name']}")
                report_lines.append(f"  Price: ₹{lowest['price']:,.0f}")
                report_lines.append(f"  Source: {lowest['source']}")
                if lowest.get('url'):
                    report_lines.append(f"  URL: {lowest['url']}")
            
            # Best rated
            if stats.get('best_rated'):
                best = stats['best_rated']
                report_lines.append("\n⭐ Best Rated Product:")
                report_lines.append(f"  Product: {best['product_name']}")
                report_lines.append(f"  Rating: {best['rating']}/5.0")
                report_lines.append(f"  Price: ₹{best['price']:,.0f}")
                report_lines.append(f"  Source: {best['source']}")
            
            # Average rating by source
            if 'rating' in df_to_analyze.columns and 'source' in df_to_analyze.columns:
                report_lines.append("\nAverage Rating by Source:")
                for source in df_to_analyze['source'].unique():
                    source_df = df_to_analyze[df_to_analyze['source'] == source]
                    avg_rating = source_df['rating'].mean()
                    if pd.notna(avg_rating):
                        report_lines.append(f"  - {source}: {avg_rating:.2f}/5.0")
            
            report_lines.append("")
            report_lines.append("=" * 70)
            report_lines.append(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            report_lines.append("=" * 70)
            
            report = "\n".join(report_lines)
            logger.info("✓ Summary report generated")
            return report
            
        except Exception as e:
            logger.error(f"Error generating summary report: {e}", exc_info=True)
            return f"Error generating summary report: {str(e)}"
    
    # Backward compatibility methods
    def combine_data(self, products_list: List[List[Dict]]) -> pd.DataFrame:
        """Backward compatibility wrapper for combine_scraped_data."""
        all_products = []
        for products in products_list:
            all_products.extend(products)
        
        # Split by source for combine_scraped_data
        amazon_data = [p for p in all_products if p.get('source', '').lower() in ['amazon', 'amazon india']]
        flipkart_data = [p for p in all_products if p.get('source', '').lower() == 'flipkart']
        
        return self.combine_scraped_data(amazon_data, flipkart_data)
    
    def filter_data(self, min_price: float = 0, max_price: float = float('inf'), 
                   min_rating: float = 0) -> pd.DataFrame:
        """Filter products based on criteria."""
        if self.df is None or self.df.empty:
            logger.warning("No data to filter")
            return pd.DataFrame()
        
        filtered_df = self.df.copy()
        
        if 'price' in filtered_df.columns:
            filtered_df = filtered_df[
                (filtered_df['price'] >= min_price) & 
                (filtered_df['price'] <= max_price)
            ]
        
        if 'rating' in filtered_df.columns and min_rating > 0:
            filtered_df = filtered_df[
                (filtered_df['rating'].notna()) & 
                (filtered_df['rating'] >= min_rating)
            ]
        
        logger.info(f"Filtered to {len(filtered_df)} products")
        return filtered_df
    
    def sort_data(self, by: str = 'price', ascending: bool = True) -> pd.DataFrame:
        """Sort products DataFrame."""
        if self.df is None or self.df.empty:
            logger.warning("No data to sort")
            return pd.DataFrame()
        
        if by not in self.df.columns:
            logger.warning(f"Column '{by}' not found, sorting by price")
            by = 'price'
        
        sorted_df = self.df.sort_values(by=by, ascending=ascending, na_position='last')
        logger.info(f"Sorted {len(sorted_df)} products by {by}")
        return sorted_df
    
    def get_summary_stats(self) -> Dict:
        """Get summary statistics (backward compatibility)."""
        return self.calculate_price_statistics()
